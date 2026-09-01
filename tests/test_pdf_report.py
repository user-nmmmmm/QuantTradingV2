import tempfile
import unittest
from pathlib import Path

import pandas as pd

from backtest.reporting.render.pdf import (
    calculate_active_risk_metrics,
    calculate_portfolio_risk_metrics,
)
from backtest.reporting import ReportGenerator


class TestPdfReport(unittest.TestCase):
    def setUp(self):
        index = pd.date_range("2025-01-01", periods=180, freq="D", tz="UTC")
        returns = pd.Series(([0.003, -0.002, 0.001, 0.004, -0.001] * 36), index=index)
        self.equity = pd.DataFrame({"equity": 10000 * (1 + returns).cumprod()}, index=index)
        self.benchmark = 10000 * (1 + returns * 0.55).cumprod()

    def test_risk_metrics_are_finite_and_tail_is_ordered(self):
        result = calculate_portfolio_risk_metrics(self.equity["equity"])
        self.assertEqual(result["status"], "ok")
        self.assertGreater(result["annualized_volatility"], 0)
        self.assertLessEqual(result["cvar_95_period"], result["var_95_period"])

    def test_active_metrics_include_alpha_beta_and_information_ratio(self):
        result = calculate_active_risk_metrics(self.equity["equity"], self.benchmark)
        self.assertEqual(result["status"], "ok")
        self.assertIn("beta", result)
        self.assertIn("information_ratio", result)

    def test_compact_profile_keeps_pdf_dashboard_and_core_csvs(self):
        with tempfile.TemporaryDirectory() as directory:
            ReportGenerator(directory).generate(
                [], self.equity, benchmark_curve=self.benchmark,
                report_profile="compact",
            )
            names = {item.name for item in Path(directory).iterdir()}
            self.assertEqual(
                names,
                {"equity.csv", "benchmark.csv", "report.pdf", "dashboard.png"},
            )
            self.assertGreater((Path(directory) / "report.pdf").stat().st_size, 10_000)

    def test_workbook_profile_consolidates_all_outputs(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            ReportGenerator(directory).generate(
                [], self.equity, benchmark_curve=self.benchmark,
                report_profile="workbook",
            )
            names = {item.name for item in Path(directory).iterdir()}
            self.assertEqual(names, {"backtest_report.xlsx"})
            workbook = load_workbook(Path(directory) / "backtest_report.xlsx", read_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["Dashboard", "Equity", "Monthly Returns", "Drawdowns",
                 "Trade Analysis", "Trades", "Closed Trades", "Benchmark",
                 "Data Quality"],
            )
            self.assertGreaterEqual(len(workbook["Dashboard"]._charts), 2)


if __name__ == "__main__":
    unittest.main()
