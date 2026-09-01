"""Single-file Excel report for backtest research runs."""

from __future__ import annotations

import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import AreaChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from backtest.pdf_report import (
    calculate_active_risk_metrics,
    calculate_portfolio_risk_metrics,
)
from core.metrics import monthly_returns


NAVY = "17324D"
BLUE = "2F75B5"
PALE_BLUE = "EAF2F8"
PALE_GRAY = "F4F6F8"
WHITE = "FFFFFF"
RED = "C00000"
GREEN = "00875A"
THIN_GRAY = Side(style="thin", color="D9E1E8")


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return None if not np.isfinite(value) else float(value)
    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            value = value.tz_convert("UTC").tz_localize(None)
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _title(ws, text: str, end_column: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=18)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30


def _section(ws, row: int, column: int, title: str, width: int = 2) -> None:
    for col in range(column, column + width):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.border = Border(bottom=THIN_GRAY)
    ws.cell(row, column, title)


def _write_metric_block(
    ws, row: int, column: int, title: str,
    values: Mapping[str, Any], labels: Mapping[str, str], percent_keys: set[str],
) -> None:
    _section(ws, row, column, title)
    current = row + 1
    for key, value in values.items():
        if key in {"status", "sample_size"}:
            continue
        label_cell, value_cell = ws.cell(current, column), ws.cell(current, column + 1)
        label_cell.value = labels.get(key, key)
        value_cell.value = _excel_value(value)
        label_cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        label_cell.font = Font(color=NAVY)
        for cell in (label_cell, value_cell):
            cell.border = Border(bottom=THIN_GRAY)
        if key in percent_keys:
            value_cell.number_format = "0.00%"
        elif isinstance(value, (float, np.floating)):
            value_cell.number_format = "#,##0.000"
        elif isinstance(value, (int, np.integer)):
            value_cell.number_format = "#,##0"
        current += 1
    if current == row + 1:
        ws.cell(current, column, "Status")
        ws.cell(current, column + 1, str(values.get("status", "not available")))


def _write_frame(ws, frame: pd.DataFrame, table_name: str) -> None:
    frame = frame.copy()
    ws.append([str(column) for column in frame.columns])
    for row in frame.itertuples(index=False, name=None):
        ws.append([_excel_value(value) for value in row])
    header = ws[1]
    for cell in header:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)
    for column_cells in ws.columns:
        length = min(42, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        ws.column_dimensions[column_cells[0].column_letter].width = length
    ws.sheet_view.showGridLines = False


def _closed_trade_frame(closed_trades: Sequence[Mapping[str, Any]]) -> pd.DataFrame:
    if not closed_trades:
        return pd.DataFrame(columns=["status"])
    columns = sorted({key for trade in closed_trades for key in trade})
    return pd.DataFrame([{key: trade.get(key) for key in columns} for trade in closed_trades])


def _flatten_quality(value: Any, prefix: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if isinstance(value, Mapping):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten_quality(child, child_prefix))
    elif isinstance(value, (list, tuple)):
        rows.append({"check": prefix, "value": json.dumps(value, ensure_ascii=False, default=str)})
    else:
        rows.append({"check": prefix, "value": _excel_value(value)})
    return rows


def write_workbook_report(
    output_dir: str | Path,
    metrics: Mapping[str, Any],
    equity_curve: pd.DataFrame,
    raw_trades: Sequence[Mapping[str, Any]],
    closed_trades: Sequence[Mapping[str, Any]],
    benchmark_curve: pd.Series | None = None,
    metadata: Mapping[str, Any] | None = None,
    data_quality: Mapping[str, Any] | None = None,
) -> Path:
    """Create one editable workbook containing summaries, tables and raw data."""
    root = Path(output_dir)
    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"
    dashboard.sheet_view.showGridLines = False
    _title(dashboard, "Backtest Research Workbook", 18)
    visible_meta = {
        key: value for key, value in (metadata or {}).items()
        if key in {"Days", "Start", "End", "Capital", "Symbols", "Source", "BenchmarkMode", "Timeframe"}
    }
    dashboard["A2"] = " | ".join(f"{k}: {v}" for k, v in visible_meta.items())
    dashboard["A2"].font = Font(color="667788", size=9)
    dashboard["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    dashboard.merge_cells("A2:H2")
    dashboard.row_dimensions[2].height = 28

    risk = calculate_portfolio_risk_metrics(equity_curve["equity"])
    active = calculate_active_risk_metrics(equity_curve["equity"], benchmark_curve)
    labels = {
        "TotalReturn": "Total return", "CAGR": "CAGR", "MaxDrawdownPct": "Max drawdown",
        "SharpeRatio": "Sharpe", "TotalTrades": "Closed trades", "WinRate": "Win rate",
        "ProfitFactor": "Profit factor", "NetPnL": "Net PnL", "EndEquity": "End equity",
        "annualized_return_arithmetic": "Annual return (mean)", "annualized_volatility": "Annual volatility",
        "downside_deviation": "Downside deviation", "sortino_ratio": "Sortino", "calmar_ratio": "Calmar",
        "var_95_period": "Historical VaR 95%", "cvar_95_period": "Historical CVaR 95%",
        "return_skewness": "Return skewness", "return_excess_kurtosis": "Excess kurtosis",
        "positive_period_ratio": "Positive periods", "positive_month_ratio": "Positive months",
        "best_month": "Best month", "worst_month": "Worst month", "beta": "Beta",
        "annualized_alpha": "Annual alpha", "tracking_error": "Tracking error",
        "information_ratio": "Information ratio", "return_correlation": "Correlation",
        "up_capture": "Up capture", "down_capture": "Down capture",
    }
    percent_keys = {
        "TotalReturn", "CAGR", "MaxDrawdownPct", "WinRate", "annualized_return_arithmetic",
        "annualized_volatility", "downside_deviation", "var_95_period", "cvar_95_period",
        "positive_period_ratio", "positive_month_ratio", "best_month", "worst_month",
        "annualized_alpha", "tracking_error",
    }
    core_keys = ("TotalReturn", "CAGR", "MaxDrawdownPct", "SharpeRatio", "TotalTrades", "WinRate", "ProfitFactor", "NetPnL", "EndEquity")
    _write_metric_block(dashboard, 4, 1, "Performance", {key: metrics.get(key) for key in core_keys}, labels, percent_keys)
    _write_metric_block(dashboard, 4, 4, "Portfolio risk", risk, labels, percent_keys)
    _write_metric_block(dashboard, 4, 7, "Benchmark-relative", active, labels, percent_keys)
    for column in ("A", "D", "G"):
        dashboard.column_dimensions[column].width = 25
    for column in ("B", "E", "H"):
        dashboard.column_dimensions[column].width = 15
    dashboard.freeze_panes = "A4"

    equity = equity_curve.copy()
    equity.index.name = "timestamp"
    equity = equity.reset_index()
    if benchmark_curve is not None:
        aligned = benchmark_curve.reindex(equity_curve.index)
        equity["benchmark"] = aligned.to_numpy()
    else:
        equity["benchmark"] = np.nan
    peak = equity["equity"].cummax().replace(0, np.nan)
    equity["drawdown"] = (equity["equity"] - peak) / peak
    equity["period_return"] = equity["equity"].pct_change(fill_method=None)
    equity["date_label"] = pd.to_datetime(equity["timestamp"]).dt.strftime("%Y-%m-%d")
    equity["chart_strategy"] = equity["equity"]
    equity["chart_benchmark"] = equity["benchmark"]
    equity["chart_drawdown"] = equity["drawdown"]
    equity_ws = wb.create_sheet("Equity")
    _write_frame(equity_ws, equity, "EquityTable")
    for cell in equity_ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm"
    for col in ("B", "C", "D"):
        for cell in equity_ws[col][1:]:
            cell.number_format = "#,##0.00"
    for col in ("E", "F", "J"):
        for cell in equity_ws[col][1:]:
            cell.number_format = "0.00%"

    max_row = equity_ws.max_row
    if max_row >= 3:
        line = LineChart()
        line.title = "Strategy vs Benchmark"
        line.y_axis.title = "Equity"
        line.x_axis.title = "Date"
        line.add_data(Reference(equity_ws, min_col=8, max_col=9, min_row=1, max_row=max_row), titles_from_data=True)
        line.set_categories(Reference(equity_ws, min_col=7, min_row=2, max_row=max_row))
        line.series[0].tx = SeriesLabel(v="Strategy")
        line.series[1].tx = SeriesLabel(v="Benchmark")
        line.height, line.width = 8.0, 17.0
        dashboard.add_chart(line, "J4")
        area = AreaChart()
        area.title = "Underwater Drawdown"
        area.y_axis.title = "Drawdown"
        area.add_data(Reference(equity_ws, min_col=10, min_row=1, max_row=max_row), titles_from_data=True)
        area.set_categories(Reference(equity_ws, min_col=7, min_row=2, max_row=max_row))
        area.series[0].tx = SeriesLabel(v="Drawdown")
        area.y_axis.numFmt = "0.0%"
        area.height, area.width = 7.5, 17.0
        dashboard.add_chart(area, "J20")

    monthlies = monthly_returns(equity_curve["equity"])
    if len(monthlies):
        calendar = monthlies.to_frame("return")
        calendar["year"], calendar["month"] = calendar.index.year, calendar.index.month
        calendar = calendar.pivot(index="year", columns="month", values="return").reindex(columns=range(1, 13))
        calendar.columns = [f"{month:02d}" for month in range(1, 13)]
        calendar.index.name = "year"
        calendar = calendar.reset_index()
    else:
        calendar = pd.DataFrame(columns=["year"] + [f"{month:02d}" for month in range(1, 13)])
    monthly_ws = wb.create_sheet("Monthly Returns")
    _write_frame(monthly_ws, calendar, "MonthlyReturnsTable")
    if monthly_ws.max_row >= 2:
        monthly_ws.conditional_formatting.add(
            f"B2:M{monthly_ws.max_row}",
            ColorScaleRule(start_type="min", start_color="F8696B", mid_type="num", mid_value=0, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
        )
        for row in monthly_ws.iter_rows(min_row=2, min_col=2, max_col=13):
            for cell in row:
                cell.number_format = "0.0%"

    events = (metrics.get("ExtendedAnalytics") or {}).get("drawdown_events") or []
    drawdowns = pd.DataFrame(events)
    if not drawdowns.empty and "depth_pct" in drawdowns:
        drawdowns = drawdowns.sort_values("depth_pct")
    drawdown_ws = wb.create_sheet("Drawdowns")
    _write_frame(drawdown_ws, drawdowns if not drawdowns.empty else pd.DataFrame(columns=["status"]), "DrawdownTable")
    if "depth_pct" in drawdowns:
        depth_col = list(drawdowns.columns).index("depth_pct") + 1
        for cell in drawdown_ws.iter_cols(min_col=depth_col, max_col=depth_col, min_row=2, max_row=drawdown_ws.max_row):
            for item in cell:
                item.number_format = "0.00%"

    quality = (metrics.get("ExtendedAnalytics") or {}).get("trade_quality") or {}
    breakdown_rows = []
    for dimension in ("by_strategy", "by_symbol"):
        for name, values in (quality.get(dimension) or {}).items():
            breakdown_rows.append({"dimension": dimension[3:], "name": name, **values})
    trade_analysis_ws = wb.create_sheet("Trade Analysis")
    _write_frame(trade_analysis_ws, pd.DataFrame(breakdown_rows) if breakdown_rows else pd.DataFrame(columns=["status"]), "TradeAnalysisTable")

    trades_ws = wb.create_sheet("Trades")
    raw_frame = pd.DataFrame(raw_trades)
    _write_frame(trades_ws, raw_frame if not raw_frame.empty else pd.DataFrame(columns=["status"]), "TradesTable")
    closed_ws = wb.create_sheet("Closed Trades")
    _write_frame(closed_ws, _closed_trade_frame(closed_trades), "ClosedTradesTable")

    benchmark_ws = wb.create_sheet("Benchmark")
    benchmark_frame = (
        pd.DataFrame({"timestamp": benchmark_curve.index, "benchmark": benchmark_curve.to_numpy()})
        if benchmark_curve is not None else pd.DataFrame(columns=["timestamp", "benchmark"])
    )
    _write_frame(benchmark_ws, benchmark_frame, "BenchmarkTable")

    quality_ws = wb.create_sheet("Data Quality")
    quality_rows = _flatten_quality(data_quality or {})
    _write_frame(quality_ws, pd.DataFrame(quality_rows) if quality_rows else pd.DataFrame(columns=["check", "value"]), "DataQualityTable")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    target = root / "backtest_report.xlsx"
    wb.save(target)
    return target


__all__ = ["write_workbook_report"]
